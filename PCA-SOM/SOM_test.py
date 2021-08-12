from sklearn import datasets
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import pandas as pd
from sklearn import datasets

from minisom import MiniSom

import math
import xlrd
from icecream import ic
from tqdm import tqdm

from openpyxl import load_workbook
import openpyxl



# 传入数据成numpy的矩阵格式
def loaddata(datafile, num_name):
    df = pd.read_excel(datafile, sheet_name=num_name, index_col=0)

    return df




# 导入原始数据
# 1.导入训练和测试数据集
datafile = "GD523INITA1_0.xls"  # 原始数据

# 2.导入标签数据
y = pd.DataFrame(pd.read_csv('label.csv'))
# ic(y)
yy = []
# print('label:', y['label'])
for iy in range(y.shape[0]):
    Uy = y.iloc[iy, 0]
    yy.append(int(Uy))
# print('yyyyyyyyyyyyyy:', yy)

y = yy


feature_names = pd.DataFrame(pd.read_excel(datafile, index_col=0)).columns
class_names = [0, 1]
# ic(feature_names, class_names)
feat = []
for tz in range(feature_names.shape[0]):
    tezh = feature_names[tz]

    feat.append(tezh)

feature_names = feat

print('特征名称:', feature_names)

# 按照每一个数据Sheet读取每一层的数据
# for i_c in range
workbook = xlrd.open_workbook(datafile)
sheets = workbook.sheet_names()
SheetNames = []   # 读取原始数据的数据表sheet名
for sheetname in tqdm(sheets):
    print("表格的名称是：", sheetname)
    SheetNames.append(sheetname)

print('原始数据表的表单名称为：', SheetNames)
num_n = pd.DataFrame(SheetNames).shape[0]  # 获取表单的个数
print('表单的个数为：', num_n)

# 设置空数据表1
dff = pd.DataFrame(columns=["title", "content"])
dff.to_excel('SOM_Result.xlsx')

## 设置空数据表2
dff2 = pd.DataFrame(columns=["title1", "content1"])
dff2.to_excel('SOM_label_result.xlsx')


for i_c in range(num_n):
    if i_c < num_n:
        print('程序目前处在第%r层数.' % SheetNames[i_c])
        XMat = loaddata(datafile, num_name=SheetNames[i_c])  # 返回得到浮点型矩阵

        X =  XMat.values    # 将DataFrame格式改为np.array




        # 划分训练集、测试集  7:3
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=0)


        N = X_train.shape[0]  #样本数量
        M = X_train.shape[1]  #维度/特征数量


        '''
        设置超参数
        '''
        size = math.ceil(np.sqrt(5 * np.sqrt(N)))  # 经验公式：决定输出层尺寸
        print("训练样本个数:{}  测试样本个数:{}".format(N, X_test.shape[0]))
        print("输出网格最佳边长为:", size)

        # ic(size)

        max_iter = 8000  # 迭代次数

        # Initialization and training
        som = MiniSom(size, size, M, sigma=3, learning_rate=0.5,
                      neighborhood_function='bubble')    # Neighborhood_function可选的设置有'gaussian'、'mexican_hat'、'bubble'. 调参的时候可以都试一遍，看效果

        '''
        初始化权值，有2个API
        '''
        #som.random_weights_init(X_train)
        som.pca_weights_init(X_train)

        som.train_batch(X_train, max_iter, verbose=False)
        #som.train_random(X_train, max_iter, verbose=False)

        winmap = som.labels_map(X_train, y_train)

        def classify(som,data,winmap):
            from numpy import sum as npsum
            default_class = npsum(list(winmap.values())).most_common()[0][0]
            result = []
            for d in data:
                win_position = som.winner(d)
                if win_position in winmap:
                    result.append(winmap[win_position].most_common()[0][0])
                else:
                    result.append(default_class)

            print('输出result结果：', result)
            return result

        # 输出混淆矩阵
        y_pred = classify(som, X_test, winmap)
        print(classification_report(y_test, np.array(y_pred)))

        # U-Matrix
        heatmap = som.distance_map()  #生成U-Matrix
        plt.imshow(heatmap, cmap='bone_r')      #miniSom案例中用的pcolor函数,需要调整坐标
        plt.colorbar()

        plt.figure(figsize=(9, 9))
        # 背景上画U-Matrix
        heatmap = som.distance_map()
        plt.pcolor(heatmap, cmap='bone_r')  # plotting the distance map as background

        # 定义不同标签的图案标记
        markers = ['o', 's', 'D']
        colors = ['C0', 'C1', 'C2']
        category_color = {'setosa': 'C0',
                          'versicolor': 'C1',
                          'virginica': 'C2'}

        for cnt, xx in enumerate(X_train):
            w = som.winner(xx)  # getting the winner
            # 在样本Heat的地方画上标记
            plt.plot(w[0]+.5, w[1]+.5, markers[y_train[cnt]], markerfacecolor='None',
                     markeredgecolor=colors[y_train[cnt]], markersize=12, markeredgewidth=2)
        plt.axis([0, size, 0, size])
        ax = plt.gca()
        ax.invert_yaxis() #颠倒y轴方向
        legend_elements = [Patch(facecolor=clr,
                                 edgecolor='w',
                                 label=l) for l, clr in category_color.items()]
        plt.legend(handles=legend_elements, loc='center left', bbox_to_anchor=(1, .95))
        plt.show()


        label_name_map_number = {"setosa":0,"versicolor":1,"virginica":2}

        from matplotlib.gridspec import GridSpec
        plt.figure(figsize=(9, 9))
        the_grid = GridSpec(size, size)
        for position in winmap.keys():
            label_fracs = [winmap[position][label] for label in [0,1,2]]
            plt.subplot(the_grid[position[1], position[0]], aspect=1)
            patches, texts = plt.pie(label_fracs)
            plt.text(position[0]/100, position[1]/100,  str(len(list(winmap[position].elements()))),
                      color='black', fontdict={'weight': 'bold',  'size': 15},
                      va='center',ha='center')
        plt.legend(patches, class_names, loc='center right', bbox_to_anchor=(-1,9), ncol=3)
        plt.show()

        plt.figure(figsize=(10, 10))
        for i, f in enumerate(feature_names):
            plt.subplot(6, 6, i+1)
            plt.title(f)
            W = som.get_weights()
            plt.imshow(W[:,:,i], cmap='coolwarm')
            plt.colorbar()
            plt.xticks(np.arange(size+1))
            plt.yticks(np.arange(size+1))
        #plt.tight_layout()
        plt.show()

        # 保存result——label
        print('开始SOM标签Result保存！')
        df_winmap = pd.DataFrame.from_dict(winmap, orient='index')
        ic(df_winmap)
        writer1 = pd.ExcelWriter('SOM_label_result.xlsx', engine='openpyxl')
        book1 = load_workbook(writer1.path)
        writer1.book = book1
        df_winmap.to_excel(excel_writer=writer1, sheet_name=str(SheetNames[i_c]))
        writer1.save()
        writer1.close()
        print('SOM标签Result保存结束！')
        # ic(winmap)

        # 保存result_data
        print('开始SOM最终Result坐标保存！')
        winner = som.win_map(X_train, return_indices=True)

        # my_list = [winner]

        my_df = pd.DataFrame.from_dict(winner, orient='index')
        ic(my_df)
        writer = pd.ExcelWriter('SOM_Result.xlsx', engine='openpyxl')
        book = load_workbook(writer.path)
        writer.book = book
        my_df.to_excel(excel_writer=writer, sheet_name=str(SheetNames[i_c]))

        writer.save()
        writer.close()
        print('SOM最终Result坐标保存结束！')


# 删除空表头Sheet1
# 执行删除操作：
sheet_name1 = 'Sheet1'
# 载入工作簿
wb = openpyxl.load_workbook('SOM_Result.xlsx')
#
# 删除目标Sheet
ws = wb[sheet_name1]
wb.remove(ws)

# 保存已做删除处理的工作簿
wb.save('SOM_Result.xlsx')
print('工作簿1处理完毕！')

# 删除空表头Sheet2
# 执行删除操作：
sheet_name2 = 'Sheet1'
# 载入工作簿
wb2 = openpyxl.load_workbook('SOM_label_result.xlsx')
#
# 删除目标Sheet
ws2 = wb2[sheet_name2]
wb2.remove(ws2)

# 保存已做删除处理的工作簿
wb2.save('SOM_label_result.xlsx')
print('工作簿2处理完毕！')



print('聚类结束！')
print('Success!!!')

